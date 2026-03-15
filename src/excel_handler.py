"""Excel file handling for reading addresses and writing results."""

import os
from pathlib import Path
from typing import Optional, Dict
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image
import sys

from config import EXCEL_COLUMNS, IMAGE_WIDTH
from console_helper import console


class ExcelHandler:
    """Handles reading from and writing to Excel files."""

    def __init__(self, file_path: str):
        """
        Initialize Excel handler.

        Args:
            file_path: Path to the Excel file
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.temp_image_files = []  # Track temporary resized images for cleanup

    def open(self) -> bool:
        """
        Open the Excel file.

        Returns:
            True if successful, False otherwise
        """
        try:
            if not os.path.exists(self.file_path):
                console.print(f"[red]Error: File not found: {self.file_path}[/red]")
                return False

            self.workbook = load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            console.print(f"[green]OK[/green] Opened Excel file: {self.file_path}")
            return True
        except Exception as e:
            console.print(f"[red]Error opening Excel file: {e}[/red]")
            return False

    def read_id(self, row: int) -> Optional[str]:
        """
        Read ID from the specified row.

        Args:
            row: Row number (1-indexed)

        Returns:
            ID string or None if empty
        """
        try:
            cell = self.worksheet[f"{EXCEL_COLUMNS['ID']}{row}"]
            val = cell.value

            if val is not None and str(val).strip():
                return str(val).strip()
            return None
        except Exception as e:
            console.print(f"[yellow]Warning: Error reading ID row {row}: {e}[/yellow]")
            return None

    def read_pnu(self, row: int) -> Optional[str]:
        """Read PNU from the specified row."""
        try:
            cell = self.worksheet[f"{EXCEL_COLUMNS['PNU']}{row}"]
            val = cell.value

            if val is not None and str(val).strip():
                return str(val).strip()
            return None
        except Exception as e:
            console.print(f"[yellow]Warning: Error reading PNU row {row}: {e}[/yellow]")
            return None

    def read_address(self, row: int) -> Optional[str]:
        """
        Read address from the specified row.

        Args:
            row: Row number (1-indexed)

        Returns:
            Address string or None if empty
        """
        try:
            cell = self.worksheet[f"{EXCEL_COLUMNS['ADDRESS_INPUT']}{row}"]
            address = cell.value

            if address and str(address).strip():
                return str(address).strip()
            return None
        except Exception as e:
            console.print(f"[yellow]Warning: Error reading row {row}: {e}[/yellow]")
            return None

    @staticmethod
    def create_template(filepath: str) -> bool:
        """Create a new Excel template file."""
        try:
            from openpyxl import Workbook
            from config import TEMPLATE_HEADERS
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Address List"
            
            # Write headers
            for col_idx, header in enumerate(TEMPLATE_HEADERS, 1):
                ws.cell(row=1, column=col_idx, value=header)
                
            wb.save(filepath)
            return True
        except Exception as e:
            console.print(f"[red]Error creating template: {e}[/red]")
            return False

    def write_data(self, row: int, data: Dict[str, str]) -> bool:
        """
        Write scraped data to the Excel file.

        Args:
            row: Row number (1-indexed)
            data: Dictionary with keys matching config SELECTORS

        Returns:
            True if successful, False otherwise
        """
        try:
            # Write text data
            if "id" in data:
                self.worksheet[f"{EXCEL_COLUMNS['ID']}{row}"] = data["id"]
            if "present_class" in data:
                self.worksheet[f"{EXCEL_COLUMNS['PRESENT_CLASS']}{row}"] = data["present_class"]
            if "present_area" in data:
                self.worksheet[f"{EXCEL_COLUMNS['PRESENT_AREA']}{row}"] = data["present_area"]
            if "jiga" in data:
                self.worksheet[f"{EXCEL_COLUMNS['JIGA']}{row}"] = data["jiga"]
            if "jiga_year" in data:
                self.worksheet[f"{EXCEL_COLUMNS['JIGA_YEAR']}{row}"] = data["jiga_year"]
            if "present_mark1" in data:
                self.worksheet[f"{EXCEL_COLUMNS['PRESENT_MARK1']}{row}"] = data["present_mark1"]
            if "present_mark2" in data:
                self.worksheet[f"{EXCEL_COLUMNS['PRESENT_MARK2']}{row}"] = data["present_mark2"]
            if "present_mark3" in data:
                self.worksheet[f"{EXCEL_COLUMNS['PRESENT_MARK3']}{row}"] = data["present_mark3"]
            if "present_mark_combined" in data:
                self.worksheet[f"{EXCEL_COLUMNS['PRESENT_MARK_COMBINED']}{row}"] = data["present_mark_combined"]
            
            # Write status/error info
            if "result" in data:
                self.worksheet[f"{EXCEL_COLUMNS['RESULT']}{row}"] = data["result"]
            if "details" in data:
                self.worksheet[f"{EXCEL_COLUMNS['DETAILS']}{row}"] = data["details"]
            if "pnu" in data:
                self.worksheet[f"{EXCEL_COLUMNS['PNU']}{row}"] = data["pnu"]
            if "image_status" in data:
                self.worksheet[f"{EXCEL_COLUMNS['IMAGE_STATUS']}{row}"] = data["image_status"]

            return True
        except Exception as e:
            console.print(f"[yellow]Warning: Error writing data to row {row}: {e}[/yellow]")
            return False

    def insert_image(self, row: int, image_path: str) -> bool:
        """
        Insert image into the Excel file.

        Args:
            row: Row number (1-indexed)
            image_path: Path to the image file

        Returns:
            True if successful, False otherwise
        """
        try:
            if not os.path.exists(image_path):
                console.print(f"[yellow]Warning: Image not found: {image_path}[/yellow]")
                return False

            # Use original image without resizing
            img = Image.open(image_path)
            
            # Convert to RGB if needed (for PNG with transparency)
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                img = background

            # Save to temporary file in temp_images directory (as PNG)
            from config import TEMP_IMAGE_DIR
            temp_path = os.path.join(TEMP_IMAGE_DIR, f"excel_original_row{row}.png")
            
            img.save(temp_path, format='PNG')
            self.temp_image_files.append(temp_path)

            # Insert into Excel from file
            xl_img = XLImage(temp_path)
            cell_ref = f"{EXCEL_COLUMNS['IMAGE']}{row}"
            self.worksheet.add_image(xl_img, cell_ref)
            
            # Adjust row height to match image height
            # Excel row height is in points. 1 pixel approx 0.75 points
            # We add a little padding
            row_height = (img.height * 0.75) + 10
            self.worksheet.row_dimensions[row].height = row_height

            return True
        except Exception as e:
            console.print(f"[yellow]Warning: Error inserting image to row {row}: {e}[/yellow]")
            return False

    def save(self) -> bool:
        """
        Save the Excel file.

        Returns:
            True if successful, False otherwise
        """
        try:
            self.workbook.save(self.file_path)
            # Don't delete temp files yet - openpyxl reads them on every save()
            # Files will be cleaned up in close() or when program exits
            return True
        except Exception as e:
            console.print(f"[red]Error saving Excel file: {e}[/red]")
            return False

    def close(self):
        """Close the Excel file."""
        if self.workbook:
            self.workbook.close()

        # Clean up temporary image files after closing
        for temp_file in self.temp_image_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except Exception:
                pass  # Ignore cleanup errors
        self.temp_image_files.clear()

    def __enter__(self):
        """Context manager entry."""
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        self.close()
