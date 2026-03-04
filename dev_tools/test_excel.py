"""Quick test to verify Excel file can be read."""

from openpyxl import load_workbook

try:
    wb = load_workbook("example_data.xlsx")
    ws = wb.active

    print("✓ Excel file opened successfully")
    print(f"✓ Sheet name: {ws.title}")
    print(f"✓ Dimensions: {ws.dimensions}")

    print("\nData in file:")
    for row in range(1, 5):
        col_a = ws[f"A{row}"].value
        col_b = ws[f"B{row}"].value
        print(f"  Row {row}: A='{col_a}', B='{col_b}'")

    wb.close()
    print("\n✓ File is valid and readable")

except Exception as e:
    print(f"✗ Error reading file: {e}")
    import traceback
    traceback.print_exc()
