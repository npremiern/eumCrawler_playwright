"""Script to create example Excel file."""

from openpyxl import Workbook

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "부동산 정보"

# Add headers
headers = ["NO", "주소 (입력)", "주소", "분류", "면적", "지가", "표시1", "표시2", "표시3", "이미지"]
for col_idx, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_idx, value=header)

# Add example addresses
example_addresses = [
    "서울특별시 강남구 테헤란로 152",
    "서울특별시 종로구 세종대로 209",
    "부산광역시 해운대구 해운대해변로 264",
]

for row_idx, address in enumerate(example_addresses, start=2):
    ws.cell(row=row_idx, column=1, value=row_idx - 1)  # ID
    ws.cell(row=row_idx, column=2, value=address)  # Address

# Adjust column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 20

# Save file
wb.save("example_data.xlsx")
print("✓ Created example_data.xlsx")
